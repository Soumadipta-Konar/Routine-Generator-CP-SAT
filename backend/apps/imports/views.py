import os
import tempfile
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .parser import run_ingestion_pipeline

class MasterWorkbookImportView(APIView):
    """
    POST /api/v1/imports/master-workbook/
    Accepts multipart/form-data with an Excel file ('file') and populates the database.
    """
    def post(self, request):
        if 'file' not in request.FILES:
            return Response({"error": "No file uploaded. Please provide a 'file' key in form-data."}, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({"error": "Invalid file format. Only Excel (.xlsx, .xls) files are supported."}, status=status.HTTP_400_BAD_REQUEST)

        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            for chunk in excel_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            success, errors = run_ingestion_pipeline(tmp_path)
            if not success:
                is_db_err = any(e.get('sheet') == 'Database' or e.get('error_type') in ['Transaction_Aborted', 'Connection_Error'] for e in (errors or []))
                err_list = [f"[{e.get('sheet', 'Workbook')}] {e.get('description', 'Validation error')}" for e in (errors or [])]
                err_summary = " \n ".join(err_list) if err_list else "Validation failed. Please verify sheet formats."
                return Response({
                    "error": "DATABASE_ERROR" if is_db_err else "VALIDATION_FAILED",
                    "message": err_summary,
                    "details": errors
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR if is_db_err else status.HTTP_422_UNPROCESSABLE_ENTITY)

            return Response({
                "success": True,
                "message": "Master workbook successfully parsed and database populated.",
                "details": errors
            }, status=status.HTTP_200_OK)
        except ValueError as ve:
            return Response({"error": "VALIDATION_ERROR", "message": str(ve)}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
        except Exception as e:
            return Response({"error": "SERVER_ERROR", "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass


import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

class TemplateDownloadView(APIView):
    """
    GET /api/v1/imports/download-template/
    Generates and returns a clean, formatted blank Excel workbook with all 5 required sheets.
    """
    def get(self, request):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1C1917", end_color="1C1917", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        
        sheets_data = {
            "Cohorts": ["Cohort_ID", "Name", "Semester", "Size", "Department"],
            "Faculty": ["Faculty_ID", "Name", "Contact_Number", "Email", "Department", "Max_Weekly_Hours"],
            "Rooms": ["Room_ID", "Name", "Capacity", "Room_Type"],
            "Subjects": ["Subject_ID", "Code", "Name", "Subject_Type", "Is_Heavy_Cognitive", "Periods_Per_Week", "Department", "Required_Capabilities"],
            "Curriculum_Workload": ["Mapping_ID", "Cohort_ID", "Subject_ID", "Faculty_ID", "Weekly_Periods", "Smart_Class_Requirement"],
        }
        
        for sheet_name, columns in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.append(columns)
            
            # Style header row
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                
                # Auto-adjust column width
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(col_name) + 6, 16)
                
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="master_schedule_template.xlsx"'
        return response

class SampleDownloadView(APIView):
    """
    GET /api/v1/imports/download-sample/
    Returns the populated sample Excel file (master_schedule_sample.xlsx).
    """
    def get(self, request):
        sample_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), 'master_schedule_sample.xlsx')
        if os.path.exists(sample_path):
            with open(sample_path, 'rb') as f:
                content = f.read()
            response = HttpResponse(
                content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = 'attachment; filename="master_schedule_sample.xlsx"'
            return response
        else:
            return Response({"error": "Sample file not found."}, status=status.HTTP_404_NOT_FOUND)


